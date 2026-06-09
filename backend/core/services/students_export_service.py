import csv
import io

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from ..models import Estudiantes, Inscripciones
from ..tracing import trace_service_class
from .access_service import AccessControlService


@trace_service_class
class StudentsExportService:

    def __init__(self):
        self.ac = AccessControlService()

    def export_csv(self, usuario, gestion=None, grado_id=None, incluir_inactivos=False):
        if not self.ac.puede_ver_todo(usuario):
            raise PermissionError('No autorizado')
        qs = self._build_queryset(gestion, grado_id, incluir_inactivos)
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['RUDE', 'CI', 'Nombres', 'Primer Apellido', 'Segundo Apellido',
                         'Fecha Nacimiento', 'Genero', 'Estado', 'Curso', 'Gestion'])
        for e, insc in qs:
            writer.writerow([
                e.rude, e.ci, e.nombres, e.primer_apellido, e.segundo_apellido or '',
                str(e.fecha_nacimiento or ''), e.genero or '', e.estado,
                str(insc.curso) if insc else '', insc.gestion if insc else '',
            ])
        return buf.getvalue()

    def export_xlsx(self, usuario, gestion=None, grado_id=None, incluir_inactivos=False):
        if not self.ac.puede_ver_todo(usuario):
            raise PermissionError('No autorizado')
        qs = self._build_queryset(gestion, grado_id, incluir_inactivos)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estudiantes'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = openpyxl.styles.PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        headers = ['RUDE', 'CI', 'Nombres', 'Primer Apellido', 'Segundo Apellido',
                   'Fecha Nacimiento', 'Genero', 'Estado', 'Curso', 'Gestion']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row_idx, (e, insc) in enumerate(qs, 2):
            values = [
                e.rude, e.ci, e.nombres, e.primer_apellido, e.segundo_apellido or '',
                str(e.fecha_nacimiento or ''), e.genero or '', e.estado,
                str(insc.curso) if insc else '', insc.gestion if insc else '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _build_queryset(self, gestion=None, grado_id=None, incluir_inactivos=False):
        estudiantes = Estudiantes.objects.all()
        if not incluir_inactivos:
            estudiantes = estudiantes.filter(estado='activo')
        if gestion or grado_id:
            insc_filter = {}
            if gestion:
                insc_filter['gestion'] = gestion
            inscs = Inscripciones.objects.filter(**insc_filter).select_related('curso__grado', 'curso__paralelo')
            if grado_id:
                inscs = inscs.filter(curso__grado_id=grado_id)
            estudiante_ids = inscs.values_list('estudiante_id', flat=True)
            estudiantes = estudiantes.filter(id__in=estudiante_ids)
            insc_map = {i.estudiante_id: i for i in inscs}
        else:
            insc_map = {}
        result = []
        for e in estudiantes.select_related().order_by('primer_apellido', 'nombres'):
            insc = insc_map.get(e.id)
            result.append((e, insc))
        return result
